# Copyright (c) 2024, elius mgani and contributors
# For license information, please see license.txt

import frappe
from datetime import datetime
from openpyxl import load_workbook
from frappe.model.document import Document
from frappe.core.doctype.file.utils import delete_file
from frappe.utils import escape_html


def get_sheet(workbook, seq_no):
    suffix = f"({seq_no})"
    matches = [ws for ws in workbook.worksheets if ws.title.strip().endswith(suffix)]
    if len(matches) != 1:
        sheet_names = escape_html(", ".join(workbook.sheetnames))
        frappe.throw(
            f"Expected exactly one sheet ending with <b>({seq_no})</b> in the "
            f"manifest file, found {len(matches)}. Sheets found: {sheet_names}"
        )
    return matches[0]


class Manifest(Document):
    def before_save(self):
        if not self.manifest and self.is_new():
            frappe.throw("Please attach the manifest file before saving.")

        if not self.is_new() and not self.manifest:
            frappe.throw(
                "You cannot remove the manifest file once it has been attached.<br>\
                Please delete the record and create a new one."
            )

        if not self.company:
            self.company = frappe.defaults.get_user_default("Company")

    def before_submit(self):
        if not self.port:
            frappe.throw("Please fill the <b>Port</b> before submitting.")

    def on_submit(self):
        self.create_consignees()

    def on_trash(self):
        if self.manifest:
            delete_file(self.manifest)

    @frappe.whitelist()
    def extract_data_from_manifest_file(self):
        if not self.manifest:
            frappe.throw(
                frappe._(
                    "<b>Manifest File</b> is missing, Please attach the manifest file before extracting data."
                )
            )

        # ICD Code validation
        icd_code = frappe.db.get_single_value("ICD TZ Settings", "icd_code")
        if not icd_code:
            frappe.throw(
                frappe._(
                    "ICD Code is not configured. Please set it in <b>ICD TZ Settings</b> before extracting manifest data."
                )
            )
        icd_code = icd_code.strip()

        file_url = self.manifest
        file_path = frappe.get_site_path(
            "private", "files", file_url.split("/files/")[-1]
        )

        # Load the Excel file
        workbook = load_workbook(file_path, data_only=True)

        # Function to convert dates
        def convert_date(excel_date):
            if isinstance(excel_date, datetime):
                return excel_date.strftime("%Y-%m-%d")
            if isinstance(excel_date, str):
                try:
                    return datetime.strptime(excel_date, "%d/%m/%Y").strftime(
                        "%Y-%m-%d"
                    )
                except ValueError:
                    return None
            return None

        # Process the MRN Detail (1) sheet
        vessel_info_sheet = get_sheet(workbook, 1)
        vessel_info_row = next(vessel_info_sheet.iter_rows(min_row=4, values_only=True))
        self.mrn = vessel_info_row[0]
        self.vessel_name = vessel_info_row[1]
        self.call_sign = vessel_info_row[2]
        self.voyage_no = vessel_info_row[3]
        self.arrival_date = convert_date(vessel_info_row[5])
        self.tpa_uid = vessel_info_row[6]

        # Build ICD allowlist from Master BL List (4) first — col[4] is place_of_delivery
        master_bl_sheet = get_sheet(workbook, 4)
        allowed_mbl_nos = self._get_allowed_mbl_nos(master_bl_sheet, icd_code)
        if not allowed_mbl_nos:
            frappe.throw(
                f"No records found for ICD Code <b>{icd_code}</b> in this manifest file. "
                "Please verify the ICD Code in <b>ICD TZ Settings</b> matches the "
                "'Place of Delivery' values in the Master BL sheet."
            )

        # Process the Container (2) sheet
        containers_sheet = get_sheet(workbook, 2)
        self.update_container_details(containers_sheet, allowed_mbl_nos)

        # Process the HBL Container (3) sheet
        hbl_containers_sheet = get_sheet(workbook, 3)
        self.update_hbl_containers(hbl_containers_sheet, allowed_mbl_nos)

        # Process the Master BL List (4) sheet — reuse already-loaded sheet
        self.update_master_bl_details(master_bl_sheet, allowed_mbl_nos)

        # Process the House BL List (5) sheet
        house_bl_sheet = get_sheet(workbook, 5)
        self.update_house_bl_details(house_bl_sheet, allowed_mbl_nos)

        # self.save()
        return False

    def _get_allowed_mbl_nos(self, master_bl_sheet, icd_code: str) -> set:
        """Return the set of M B/L Nos whose Place of Delivery matches icd_code.

        Column layout in Master BL List (4) — 0-indexed:
          col[0] = M B/L No
          col[4] = Place of Delivery  ← holds the ICD code (e.g. WITZDL019)
        """
        allowed = set()
        for row in master_bl_sheet.iter_rows(min_row=4, values_only=True):
            if row[0] and row[4] and str(row[4]).strip() == icd_code:
                allowed.add(str(row[0]).strip())
        return allowed

    def update_container_details(self, containers_sheet, allowed_mbl_nos: set):
        self.containers = []
        for row in containers_sheet.iter_rows(min_row=4, values_only=True):
            # Skip rows whose M BL No is not in the ICD allowlist
            if not row[0] or str(row[0]).strip() not in allowed_mbl_nos:
                continue

            container = self.append("containers", {})
            container.m_bl_no = row[0]
            container.type_of_container = row[1]
            container.container_no = row[2]
            container.container_size = row[3]
            container.seal_no1 = row[4]
            container.seal_no2 = row[5]
            container.seal_no3 = row[6]
            container.freight_indicator = row[7]
            container.no_of_packages = row[8]
            container.package_unit = row[9]
            # container.volume = row[10]
            # container.volume_unit = row[11]
            container.weight = row[10]
            container.weight_unit = row[11]
            container.plug_type_of_reefer = row[12]
            container.minimum_temperature = row[13]
            container.maximum_temperature = row[14]

    def update_hbl_containers(self, hbl_containers_sheet, allowed_mbl_nos: set):
        self.hbl_containers = []
        for row in hbl_containers_sheet.iter_rows(min_row=4, values_only=True):
            # Skip rows whose M BL No is not in the ICD allowlist
            if not row[0] or str(row[0]).strip() not in allowed_mbl_nos:
                continue

            hbicontainer = self.append("hbl_containers", {})
            hbicontainer.m_bl_no = row[0]
            hbicontainer.h_bl_no = row[1]
            hbicontainer.type_of_container = row[2]
            hbicontainer.container_no = row[3]
            hbicontainer.container_size = row[4]
            hbicontainer.seal_no1 = row[5]
            hbicontainer.seal_no2 = row[6]
            hbicontainer.seal_no3 = row[7]
            hbicontainer.freight_indicator = row[8]
            hbicontainer.no_of_packages = row[9]
            hbicontainer.package_unit = row[10]
            # hbicontainer.volume = row[11]
            # hbicontainer.volume_unit = row[12]
            hbicontainer.weight = row[11]
            hbicontainer.weight_unit = row[12]
            hbicontainer.plug_type_of_reefer = row[13]
            hbicontainer.minimum_temperature = row[14]
            hbicontainer.maximum_temperature = row[15]

    def update_master_bl_details(self, master_bl_sheet, allowed_mbl_nos: set):
        self.master_bl = []
        for row in master_bl_sheet.iter_rows(min_row=4, values_only=True):
            # Skip rows whose M BL No is not in the ICD allowlist
            if not row[0] or str(row[0]).strip() not in allowed_mbl_nos:
                continue

            master_bl = self.append("master_bl", {})
            master_bl.m_bl_no = row[0]
            master_bl.cargo_classification = row[1]
            master_bl.bl_type = row[2]
            master_bl.place_of_destination = row[3]
            master_bl.place_of_delivery = row[4]
            master_bl.oil_type = row[5]
            master_bl.port_of_loading = row[6]
            master_bl.number_of_containers = row[7]
            master_bl.cargo_description = row[8]
            master_bl.number_of_package = row[9]
            master_bl.package_unit = row[10]
            master_bl.gross_weight = row[11]
            master_bl.gross_weight_unit = row[12]
            master_bl.gross_volume = row[13]
            master_bl.gross_volume_unit = row[14]
            master_bl.invoice_value = row[15]
            master_bl.invoice_currency = row[16]
            master_bl.freight_charge = row[17]
            master_bl.freight_currency = row[18]
            master_bl.imdg_code = row[19]
            master_bl.packing_type = row[20]
            master_bl.shipping_agent_code = row[21]
            master_bl.shipping_agent_name = row[22]
            master_bl.forwarder_code = row[23]
            master_bl.forwarder_name = row[24]
            master_bl.forwarder_tel = row[25]
            master_bl.exporter_name = row[26]
            master_bl.exporter_tel = row[27]
            master_bl.exporter_address = row[28]
            master_bl.exporter_tin = row[29]
            master_bl.consignee_name = row[30]
            master_bl.consignee_tel = row[31]
            master_bl.consignee_address = row[32]
            master_bl.consignee_tin = row[33]
            master_bl.notify_name = row[34]
            master_bl.notify_tel = row[35]
            master_bl.notify_address = row[36]
            master_bl.notify_tin = row[37]
            master_bl.shipping_mark = row[38]
            master_bl.net_weight = row[39]
            master_bl.net_weight_unit = row[40]

    def update_house_bl_details(self, house_bl_sheet, allowed_mbl_nos: set):
        self.house_bl = []
        for row in house_bl_sheet.iter_rows(min_row=4, values_only=True):
            # Skip rows whose M BL No is not in the ICD allowlist
            if not row[0] or str(row[0]).strip() not in allowed_mbl_nos:
                continue

            house_bl = self.append("house_bl", {})
            house_bl.m_bl_no = row[0]
            house_bl.h_bl_no = row[1]
            house_bl.cargo_classification = row[2]
            house_bl.place_of_destination = row[3]
            house_bl.net_weight = row[4]
            house_bl.net_weight_unit = row[5]
            house_bl.number_of_containers = row[6]
            house_bl.description_of_goods = row[7]
            house_bl.number_of_package = row[8]
            house_bl.package_unit = row[9]
            house_bl.gross_weight = row[10]
            house_bl.gross_weight_unit = row[11]
            house_bl.gross_volume = row[12]
            house_bl.gross_volume_unit = row[13]
            house_bl.invoice_value = row[14]
            house_bl.invoice_currency = row[15]
            house_bl.freight_charge = row[16]
            house_bl.freight_currency = row[17]
            house_bl.imdg_code = row[18]
            house_bl.packing_type = row[19]
            house_bl.shipping_agent_code = row[20]
            house_bl.shipping_agent_name = row[21]
            house_bl.forwarder_code = row[22]
            house_bl.forwarder_name = row[23]
            house_bl.exporter_name = row[24]
            house_bl.exporter_tel = row[25]
            house_bl.exporter_address = row[26]
            house_bl.exporter_tin = row[27]
            house_bl.consignee_name = row[28]
            house_bl.consignee_tel = row[29]
            house_bl.consignee_address = row[30]
            house_bl.consignee_tin = row[31]
            house_bl.notify_name = row[32]
            house_bl.notify_tel = row[33]
            house_bl.notify_address = row[34]
            house_bl.notify_tin = row[35]
            house_bl.shipping_mark = row[36]
            house_bl.oil_type = row[37]

    def create_consignees(self):
        def create_consignee(row):
            consignee = frappe.get_doc(
                {
                    "doctype": "Consignee",
                    "consignee_name": row.consignee_name,
                    "consignee_tel": row.consignee_tel,
                    "consignee_tin": row.consignee_tin,
                    "consignee_address": row.consignee_address,
                }
            )
            consignee.insert(ignore_permissions=True)

        for row in self.master_bl:
            if not row.consignee_name:
                continue

            if not frappe.db.exists("Consignee", row.consignee_name):
                create_consignee(row)

        for row in self.house_bl:
            if not row.consignee_name:
                continue

            if not frappe.db.exists("Consignee", row.consignee_name):
                create_consignee(row)

    @frappe.whitelist()
    def get_dashboard_data(self):
        total_units = len(self.containers)

        # Identify HBLs to determine LCL
        hbl_container_nos = set(
            c.container_no for c in self.hbl_containers if c.container_no
        )

        total_loose = 0
        total_empty = 0
        total_lcl = 0
        total_fcl = 0

        for c in self.containers:
            cno = (c.container_no or "").upper()

            # Empty Container
            if c.freight_indicator == "EMP" or "MTY" in cno or "EMPTY" in cno:
                total_empty += 1
            # LCL Container (has House Bills)
            elif c.container_no in hbl_container_nos:
                total_lcl += 1
            # FCL Container (Physical container, not empty, no HBLs)
            elif c.type_of_container == "C":
                total_fcl += 1
            # Loose Cargo / Vehicles / Other
            else:
                total_loose += 1

        # Count received physical units/cargo directly from Container Reception
        received_units = frappe.db.count(
            "Container Reception", {"manifest": self.name, "docstatus": 1}
        )

        pending_units = total_units - received_units
        if pending_units < 0:
            pending_units = 0

        return {
            "total_containers": total_units,
            "total_fcl": total_fcl,
            "total_lcl": total_lcl,
            "total_empty": total_empty,
            "total_loose": total_loose,
            "received_containers": received_units,
            "pending_containers": pending_units,
        }
